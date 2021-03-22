import requests
import re
import time
import datetime
import server_list
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from termcolor import colored
from multiprocessing import Manager #,Pool
import openpyxl
import parmap

print(colored("서버를 입력해주세요","cyan"))
server_name = input()
server = server_list.server(server_name)

if(server == 0):
  print(colored("존재하지 않는 서버입니다","red"))
  quit()

print(colored("길드 이름을 입력해주세요 : ","cyan"))
guild_name = input()
URL = f"https://maple.gg/guild/{server}/{guild_name}"
MEMBER_INFO = "https://maple.gg/u/"
manager = Manager()
member_information_list = manager.list()

# chrome driver setting
chrome_options = Options()
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')

# url response
driver = webdriver.Chrome(options=chrome_options)
driver.implicitly_wait(3)
driver.get(URL)
driver.find_element_by_xpath('//*[@id="btn-sync"]').click()

# alert handling
driver.switch_to.alert.accept()

html = driver.page_source
soup = BeautifulSoup(html,'html.parser')

# member_search_function
def member_search():
  frame = soup.select_one('#guild-content')
  e_characters = frame.select('section > div.mb-4.row.text-center > div')
  characters = frame.select('section > div:nth-child(5) > div')
  member_list = []

  master = True

  for ch in e_characters:
   if master == True:
     name = ch.select_one('section > div.member-grade.is-master > div > div:nth-child(1) > b > a').get_text()
     member_list.append(name)
     master = False
   else:
     name = ch.select_one('section > div.member-grade > div > div:nth-child(1) > b > a').get_text()
     member_list.append(name)

  for ch in characters:
    name = ch.select_one('section > div:nth-child(2) > b > a').get_text()
    member_list.append(name)
  
  return member_list

# memeber_information_extract_function
def extract_member_info(member_list):
  member_name = member_list
  member_url = MEMBER_INFO + str(member_name)

  result = requests.get(member_url)
  result_soup = BeautifulSoup(result.text,'html.parser')

  name = result_soup.select_one('#user-profile > section > div > div.col-lg-8 > h3 > b').get_text()
  level = result_soup.select_one('#user-profile > section > div > div.col-lg-8 > div.user-summary > ul > li:nth-child(1)').get_text()
  level = re.findall("\d+", str(level))
  user_class = result_soup.select_one('#user-profile > section > div > div.col-lg-8 > div.user-summary > ul > li:nth-child(2)').get_text()
  if result_soup.select_one('#app > div.card.border-bottom-0 > div > section > div.row.text-center > div:nth-child(1) > section > div > div > div > h1') is not None:
    muleung_max = result_soup.select_one('#app > div.card.border-bottom-0 > div > section > div.row.text-center > div:nth-child(1) > section > div > div > div > h1').get_text()
    muleung_max2 = re.findall("\d+", str(muleung_max))
  else:
    muleung_max2 = ['0']
  if result_soup.select_one('#dohangHistoryTbody > tr:nth-child(1) > td:nth-child(2) > h5') is not None:
    muleung_recent = result_soup.select_one('#dohangHistoryTbody > tr:nth-child(1) > td:nth-child(2) > h5').get_text()
    muleung_recent2 = re.findall("\d+", str(muleung_recent))
    muleung_recent_year = result_soup.select_one('#dohangHistoryTbody > tr:nth-child(1) > td:nth-child(1) > span').get_text()
    muleung_recent_month_day = result_soup.select_one('#dohangHistoryTbody > tr:nth-child(1) > td:nth-child(1) > b').get_text()
    muleung_recent_date = muleung_recent_year + ' ' + muleung_recent_month_day
  else:
    muleung_recent2 = ['0']
    muleung_recent_date = '기록없음'
    
  member_information_list.append([name]+level+[user_class]+muleung_max2+muleung_recent2+[muleung_recent_date])

  time.sleep(1)

if __name__ == '__main__':

  # MultiProcessing
  # pool = Pool(processes=4)
  parmap.map(extract_member_info,member_search(), pm_pbar=True, pm_processes=4)
  # pool.map(extract_member_info,member_search())
  # pool.close
  # pool.join

  # Save to Excel
  try: # Loading and Initiallize Data
    database = openpyxl.load_workbook(f'{guild_name} 길드원 현황.xlsx')
    main_sheet = database["길드원 목록"]
    database.remove(main_sheet)
    main_sheet = database.create_sheet("길드원 목록", 0)
    print(colored("엑셀 파일 로딩에 성공하였습니다.",'green'))
  except: # Create New File
    database = openpyxl.Workbook()
    main_sheet = database.active
    main_sheet.title = "길드원 목록"
    print(colored("엑셀 파일이 존재하지 않습니다. 새 파일을 생성합니다.",'yellow'))
  
  main_sheet.append(['닉네임','레벨','직업','무릉 최고 층수','최근 무릉 기록','최근 무릉 기록 일자'])
  for i in range(0,len(member_search()),1):
    main_sheet.append([member_information_list[i][0],member_information_list[i][1],  member_information_list[i][2],member_information_list[i][3],member_information_list[i][4] ,member_information_list[i][5]])
  database.save(f'{guild_name} 길드원 현황.xlsx')
  now = datetime.datetime.now()
  now_date = now.strftime('%Y_%m_%d')
  database.save(f'{guild_name} 길드원 현황_{now_date}.xlsx')
  
  print(colored("길드원 정보 추출에 성공하였습니다. xlsx파일로 저장합니다.",'green'))